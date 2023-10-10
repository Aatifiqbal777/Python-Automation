#read excel file
import time

import pytest
from selenium import webdriver
from openpyxl import load_workbook


# pip install pytest openpyxl pytest-excel
def get_test_data():
   workbook = load_workbook("testdata.xlsx")

   sheet = workbook.active
   data = []
   for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
       data.append(row)
   return data

@pytest.fixture
def setup_teardown():
   driver = webdriver.Chrome()
   driver.get("https://app.vwo.com/#/login")
 # Replace with your website URL
   driver.maximize_window()
   yield driver
   driver.quit()


@pytest.mark.parametrize("username, password", get_test_data())
def test_vwo_login(driver,username, password):
    email= driver.find_element(By.ID,"login-username")
    email.send_keys(username)

    pass_word =  driver.find_element(By.ID, "login-password")
    pass_word.send_keys(password)
    driver.find_element(By.ID, js-login-btn).click()
    time.sleep(3)
   print(username, password,  driver.current_url)

   if result == "fail":
       error_msg = driver.find_element(By.ID, "js-notification-box-msg")
       assert error_msg in "Your email, password, IP address or location did not match"

   else:
       assert "Dashbord" in driver.current_url