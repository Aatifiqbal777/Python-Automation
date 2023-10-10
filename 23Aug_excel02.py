#read excel file

import pytest
from selenium import webdriver
from openpyxl import load_workbook


# pip install pytest openpyxl pytest-excel
def get_test_data():
   workbook = load_workbook("testdata.xlsx")
   sheet = workbook.active
   data = []
   for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
       data.append(row)
   return data

@pytest.fixture
def setup_teardown():
   driver = webdriver.Chrome()
 # Replace with your website URL
   driver.maximize_window()
   yield driver
   driver.quit()


@pytest.mark.parametrize("username, password, get_test_data())
def test_login(username, password):
   print(username, password)

